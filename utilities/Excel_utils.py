import openpyxl


def get_rowCount(file, sheet_name):
    book = openpyxl.load_workbook(file)  # give file name from which we have to read the data
    sheet = book[sheet_name]  # give sheet name which from which we have to read the data
    return sheet.max_row  # get number of rows is given sheet


def readData(file, sheet_name, row_no, col_no):
    book = openpyxl.load_workbook(file)  # give file name from which we have to read the data
    sheet = book[sheet_name]  # give sheet name which from which we have to read the data
    return sheet.cell(row=row_no, column=col_no).value


def writeData(file, sheet_name, row_no, col_no, data):
    book = openpyxl.load_workbook(file)  # give file name from which we have to read the data
    sheet = book[sheet_name]  # give sheet name which from which we have to read the data
    sheet.cell(row=row_no, column=col_no).value = data
    book.save(file)
